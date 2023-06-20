import abc
import argparse
import datetime
import hashlib
from jinja2 import Environment, FileSystemLoader
import logging
import openpyxl
from os.path import dirname
import re
import uuid

class Renderer:
    @classmethod
    def __subclasshook__(cls, subclass):
        return (hasattr(subclass, 'render')
                and callable(subclass.render))

class JinjaTemplateRender:
    def __init__(self, searchpath=f"{dirname(__file__)}", template='ssdf_catalog.json.j2'):
        self.loader = FileSystemLoader(searchpath)
        self.environment = Environment(loader=self.loader, autoescape=True)
        try:
            self.template = self.environment.get_template(template)
            self.template.globals.update({
                'strftime': datetime.datetime.strftime,
                'uuid4': uuid.uuid4
            })
        except Exception as err:
            logging.exception(f"Template at '{template}' failed to load, must reinit!")
            logging.exception(err)
            self.template = None
    def render(self, *args, **kwargs):
        return self.template.render(*args, **kwargs)

class Transformer(metaclass=abc.ABCMeta):
    @classmethod
    def __subclasshook__(cls, subclass):
        return (hasattr(subclass, 'load')
                and callable(subclass.load)
                and hasattr(subclass, 'transform')
                and callable(subclass.transform)
                and hasattr(subclass, 'save')
                and callable(subclass.save))

class SSDFExcelOSCALTransformer:
    """A class that will transform control content from the current release of
    the SSDF document and dynamically render it into an OSCAL catalog in current
    stable release version of OSCAL.
    """
    def __init__(self, renderer=JinjaTemplateRender, template='ssdf.oscal.json.j2'):
        self.template = template
        self.source = None
        self.source_hash = None
        self.workbook = None
        self.groups = {}
        self.citations = {}
        self.controls = {}
        self.examples = {}
        self.references = {}
        self.template = template
        self.renderer = renderer(template=self.template)
        self.oscal_catalog = {}

    @staticmethod
    def escape_bad_chars(data):
        """Currently oscal-cli 0.3.x do not allow square bracket characters
        such as [ and ] without being escaped, see usnistgov/oscal-cli#80.
        """
        return data.replace('[', '(').replace(']', ')')

    def load(self, source):
        """Load the source data of the SSDF XSLX spreadsheet.
        """
        try:
            self.workbook = openpyxl.load_workbook(source)
            b  = bytearray(128*1024)
            mv = memoryview(b)
            hasher = hashlib.sha256()
            with open(source, 'rb', buffering=0) as fd:
                for n in iter(lambda : fd.readinto(mv), 0):
                    hasher.update(mv[:n])
                self.source_hash = hasher.hexdigest()

        except Exception as err:
            logging.exception(f"Loading workbook from '{source}' failed!")
            raise err

        # Source path is loadable and no err thrown, set class instance
        # to this path now it is 'valid' enough.
        self.source = source

    def transform(self):
        try:
            main_worksheet = self.workbook['SSDF']
            practices_worksheet = self.workbook['Practices']
            groups_worksheet = self.workbook['Groups']
            citation_worksheet = self.workbook['References']
        except Exception as err:
            logging.exception('Loading one of four required workbook sheets failed!')
            raise err

        try:
            # Collect groups
            # Groups is a simple one-column, four-row sheet.
            # Check that and pull in the data.
            if groups_worksheet.max_column == 1 or groups_worksheet.max_row == 4:
                for group in groups_worksheet.iter_rows(
                    groups_worksheet.min_row,
                    groups_worksheet.max_row, 
                    groups_worksheet.min_column,
                    groups_worksheet.max_column,
                    values_only=True
                ):
                    matches = re.search(
                        '(?P<group_title>^(\S+,?(\s+\S+,?)+))\s*\((?P<group_id>[a-zA-Z]{2})\)(:\s+)*(?P<group_text>(.*)$)',
                        group[0] if group and len(group) == 1 else ''
                    )
                    self.groups[matches.group('group_id')] = {
                        **matches.groupdict()
                    }
            else:
                logging.warning(f"Transforming '{source}' detected malformed groups sheet data")

            # Collect reference citations
            if citation_worksheet.max_column == 2 and citation_worksheet.max_row > 2:
                for citation in citation_worksheet.iter_rows(
                                    citation_worksheet.min_row,
                                    citation_worksheet.max_row, 
                                    citation_worksheet.min_column,
                                    citation_worksheet.max_column,
                                ):
                    # capture name in [name] to remote brackets
                    cid_match = re.search('(\w+)', citation[0].value)
                    citation_id = cid_match.group(0) if hasattr(cid_match, 'group') else citation[0].value
                    citation_link = citation[1].hyperlink.target \
                        if hasattr(citation[1], 'hyperlink') \
                        and hasattr(citation[1].hyperlink, 'target') \
                        else None
                    self.citations[citation_id] = {
                        'citation_uuid': uuid.uuid4(),
                        'citation_link': citation_link,
                        'citation_text': self.escape_bad_chars(citation[1].value.strip(citation_link).rstrip())
                            if citation_link
                            else citation[1]
                    }
            # Collect practices (i.e. controls)
            if main_worksheet.max_column == 4 and main_worksheet.max_row >= 2:
                ranges = [cell_range for cell_range in main_worksheet.merged_cells.ranges]
                for r in ranges:
                    control = self.escape_bad_chars(main_worksheet.cell(r.min_row, r.min_col).value)
                    matches = re.search('(?P<control_title>^(\S+,?(\s+\S+,?)+))\s*\((?P<control_id>[a-zA-Z]{2}.[0-9]{1})\)(:\s+)*(?P<control_text>(.*)$)', control)
                    control_id = matches.group('control_id')
                    self.controls[control_id] = {
                        'group_id': control_id[0:2],
                        'parent_control_id': None,
                        **matches.groupdict()
                    }

                # collect sub-controls (i.e. tasks)
                subcontrol_column = next(
                    idx for idx, column in enumerate(main_worksheet.columns)
                    if column[0] and column[0].value == 'Tasks'
                )
                example_column = next(
                    idx for idx, column in enumerate(main_worksheet.columns)
                    if column[0] and column[0].value == 'Notional Implementation Examples'
                )
                reference_column = next(
                    idx for idx, column in enumerate(main_worksheet.columns)
                    if column[0] and column[0].value == 'References'
                )
                for idx, row in enumerate(main_worksheet.rows):
                    # skip first row, with column labels
                    if idx == 0: continue
                    matches = re.search(
                        '(?P<control_id>^[a-zA-Z]{2}.[0-9]{1}.[0-9]{1})(:\s+)*(?P<control_text>(.*)$)',
                        row[subcontrol_column].value if row[subcontrol_column] else ''
                    )
                    control_id = matches.group('control_id')
                    parent_control_id = control_id[0:-2]
                    group_id = control_id[0:2]
                    self.controls[control_id] = {
                        'group_id': group_id,
                        'parent_control_id': parent_control_id,
                        **matches.groupdict()
                    }
                    # get examples for a subcontrol (i.e. tasks)
                    self.examples[control_id] = {}
                    for example in row[example_column].value.splitlines():
                        example_id = example[0:9]
                        example_text = self.escape_bad_chars(example[11:])
                        self.examples[control_id][example_id] = example_text

                    # get references for a subcontrol (i.e. tasks)
                    self.references[control_id] = {}
                    for reference in row[reference_column].value.splitlines():
                        reference_id, reference_text = self.escape_bad_chars(reference).split(':')
                        self.references[control_id][reference_id] = reference_text.split(',')

            else:
                logging.error(f"Transforming '{source}' detected malformed controls sheet data")

            self.oscal_catalog = self.renderer.render(
                catalog_title='NIST SP 800-218 Secure Software Development Framework 1.1 DRAFT',
                catalog_last_modified=datetime.datetime.now(datetime.timezone.utc),
                catalog_version='1.1-draft',
                oscal_version='1.0.4',
                groups=self.groups,
                controls=self.controls,
                examples=self.examples,
                references=self.references,
                citations=self.citations
            )

        except Exception as err:
            logging.exception(f"Exception in transform:\n{err}")
            raise err

    def save(self, output_file):
        try:
            output_file.write(self.oscal_catalog)
            output_file.close()
        except Exception as err:
            logging.exception(f"Error writing catalog to {output_file}")
            logging.exception(err)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert the NIST SSDF Excel spreadsheet to an OSCAL JSON catalog')
    parser.add_argument(
        '-i',
        '--input-file',
        help='Path to NIST SSDF Excel spreadsheet file',
        default='NIST.SP.800-218.SSDF-table.xlsx',
        dest='input_file',
        required=True
    )
    parser.add_argument(
        '-o',
        '--output-file',
        help='Path intended to save OSCAL JSON catalog file. If not provided print to standard out.',
        type=argparse.FileType('w'),
        dest='output_file',
        required=False
    )
    args = parser.parse_args()
    transformer = SSDFExcelOSCALTransformer()
    transformer.load(args.input_file)
    transformer.transform()

    if args.output_file:
        transformer.save(args.output_file)
    else:
        print(transformer.oscal_catalog)
